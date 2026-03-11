from __future__ import annotations

import re
from dataclasses import asdict, dataclass, field
from pathlib import Path

from django.db import transaction
from openpyxl import load_workbook

from apps.common.choices import SignatoryTypeChoices
from apps.core.models import Physician, Room, Signatory


WORKBOOK_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_MASTER_DATA_WORKBOOK = WORKBOOK_ROOT / "NAIC MEDTECH SYSTEM DATA.xlsx"

MASTER_DATA_FIELD_LABELS = {
    "requesting physician": "physicians",
    "room": "rooms",
    "medical technologist": SignatoryTypeChoices.MEDTECH,
    "pathologist": SignatoryTypeChoices.PATHOLOGIST,
}

LICENSE_PATTERN = re.compile(r"LIC\.?\s*NO\.?\s*:?\s*(.+)$", flags=re.IGNORECASE)


@dataclass
class MasterDataImportStats:
    workbook_path: str
    physician_source_count: int = 0
    room_source_count: int = 0
    signatory_source_count: int = 0
    physicians_created: int = 0
    physicians_reactivated: int = 0
    physicians_existing: int = 0
    rooms_created: int = 0
    rooms_reactivated: int = 0
    rooms_existing: int = 0
    signatories_created: int = 0
    signatories_reactivated: int = 0
    signatories_license_filled: int = 0
    signatories_existing: int = 0
    warnings: list[str] = field(default_factory=list)

    def to_dict(self):
        return asdict(self)


@dataclass
class ParsedSignatory:
    signatory_type: str
    display_name: str
    license_no: str = ""


def normalize_text(value):
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split()).strip()


def normalize_key(value):
    return normalize_text(value).casefold()


def resolve_workbook_path(workbook_path=None):
    if workbook_path:
        candidate = Path(workbook_path)
        if not candidate.is_absolute():
            candidate = Path.cwd() / candidate
        return candidate
    return DEFAULT_MASTER_DATA_WORKBOOK


def split_multiline(value):
    if not isinstance(value, str):
        return []
    return [item for item in (normalize_text(part) for part in value.splitlines()) if item]


def split_entry_blocks(value):
    if not isinstance(value, str):
        return []

    blocks = []
    current = []
    for raw_line in value.splitlines():
        line = normalize_text(raw_line)
        if line:
            current.append(line)
            continue
        if current:
            blocks.append(current)
            current = []

    if current:
        blocks.append(current)
    return blocks


def parse_signatory_block(signatory_type, block_lines):
    if not block_lines:
        return None

    display_name = block_lines[0]
    license_no = ""
    for line in block_lines[1:]:
        match = LICENSE_PATTERN.search(line)
        if match:
            license_no = normalize_text(match.group(1))
            break

    return ParsedSignatory(
        signatory_type=signatory_type,
        display_name=display_name,
        license_no=license_no,
    )


def extract_master_data(workbook_path):
    workbook_path = resolve_workbook_path(workbook_path)
    wb = load_workbook(workbook_path, data_only=True)

    physicians = {}
    rooms = {}
    signatories = {}
    warnings = []

    for ws in wb.worksheets:
        for row_idx in range(1, ws.max_row + 1):
            field_label = normalize_text(ws.cell(row_idx, 1).value)
            if not field_label:
                continue

            field_key = field_label.casefold()
            if field_key not in MASTER_DATA_FIELD_LABELS:
                continue

            raw_value = ws.cell(row_idx, 3).value
            if not raw_value:
                continue

            mapped = MASTER_DATA_FIELD_LABELS[field_key]
            if mapped == "physicians":
                for display_name in split_multiline(raw_value):
                    physicians.setdefault(normalize_key(display_name), display_name)
                continue

            if mapped == "rooms":
                for display_name in split_multiline(raw_value):
                    rooms.setdefault(normalize_key(display_name), display_name)
                continue

            for block in split_entry_blocks(raw_value):
                parsed = parse_signatory_block(mapped, block)
                if not parsed or not parsed.display_name:
                    continue

                key = (parsed.signatory_type, normalize_key(parsed.display_name))
                existing = signatories.get(key)
                if existing is None:
                    signatories[key] = parsed
                    continue

                if not existing.license_no and parsed.license_no:
                    existing.license_no = parsed.license_no
                elif (
                    existing.license_no
                    and parsed.license_no
                    and normalize_key(existing.license_no) != normalize_key(parsed.license_no)
                ):
                    warning = (
                        f"Conflicting license numbers for {parsed.display_name} "
                        f"({parsed.signatory_type}) in workbook."
                    )
                    if warning not in warnings:
                        warnings.append(warning)

    return {
        "workbook_path": str(workbook_path),
        "physicians": sorted(physicians.values(), key=str.casefold),
        "rooms": sorted(rooms.values(), key=str.casefold),
        "signatories": sorted(
            signatories.values(),
            key=lambda entry: (entry.signatory_type, entry.display_name.casefold()),
        ),
        "warnings": warnings,
    }


@transaction.atomic
def import_master_data(workbook_path=None):
    extracted = extract_master_data(workbook_path)
    stats = MasterDataImportStats(workbook_path=extracted["workbook_path"])
    stats.warnings.extend(extracted["warnings"])
    stats.physician_source_count = len(extracted["physicians"])
    stats.room_source_count = len(extracted["rooms"])
    stats.signatory_source_count = len(extracted["signatories"])

    physician_lookup = {
        normalize_key(instance.display_name): instance
        for instance in Physician.objects.all()
    }
    room_lookup = {
        normalize_key(instance.display_name): instance
        for instance in Room.objects.all()
    }
    signatory_lookup = {
        (instance.signatory_type, normalize_key(instance.display_name)): instance
        for instance in Signatory.objects.all()
    }

    for display_name in extracted["physicians"]:
        key = normalize_key(display_name)
        existing = physician_lookup.get(key)
        if existing is None:
            physician_lookup[key] = Physician.objects.create(
                display_name=display_name,
                active=True,
            )
            stats.physicians_created += 1
            continue

        if not existing.active:
            existing.active = True
            existing.save(update_fields=["active", "updated_at"])
            stats.physicians_reactivated += 1
        else:
            stats.physicians_existing += 1

    for display_name in extracted["rooms"]:
        key = normalize_key(display_name)
        existing = room_lookup.get(key)
        if existing is None:
            room_lookup[key] = Room.objects.create(
                display_name=display_name,
                active=True,
            )
            stats.rooms_created += 1
            continue

        if not existing.active:
            existing.active = True
            existing.save(update_fields=["active", "updated_at"])
            stats.rooms_reactivated += 1
        else:
            stats.rooms_existing += 1

    for parsed in extracted["signatories"]:
        key = (parsed.signatory_type, normalize_key(parsed.display_name))
        existing = signatory_lookup.get(key)
        if existing is None:
            signatory_lookup[key] = Signatory.objects.create(
                signatory_type=parsed.signatory_type,
                display_name=parsed.display_name,
                license_no=parsed.license_no,
                active=True,
            )
            stats.signatories_created += 1
            continue

        update_fields = []
        if not existing.active:
            existing.active = True
            update_fields.append("active")
            stats.signatories_reactivated += 1

        if not existing.license_no and parsed.license_no:
            existing.license_no = parsed.license_no
            update_fields.append("license_no")
            stats.signatories_license_filled += 1
        elif (
            existing.license_no
            and parsed.license_no
            and normalize_key(existing.license_no) != normalize_key(parsed.license_no)
        ):
            stats.warnings.append(
                f"Existing signatory {existing.display_name} ({existing.signatory_type}) "
                "has a different license number than the workbook. Existing value was kept."
            )

        if update_fields:
            existing.save(update_fields=[*update_fields, "updated_at"])
        else:
            stats.signatories_existing += 1

    # Keep warning order stable while removing duplicates.
    stats.warnings = list(dict.fromkeys(stats.warnings))
    return stats
