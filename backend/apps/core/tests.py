import shutil
import tempfile
from io import StringIO
from pathlib import Path

from django.core.management import call_command
from django.core.files.uploadedfile import SimpleUploadedFile
from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from django.test import override_settings
from openpyxl import Workbook

from apps.common.choices import SignatoryTypeChoices
from apps.core.master_data_import import import_master_data
from apps.core.models import Facility, LabRequest, Organization, Patient, Physician, Room, Signatory

User = get_user_model()


class RequestCreateViewTests(TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.media_root = tempfile.mkdtemp()

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.media_root, ignore_errors=True)
        super().tearDownClass()

    def setUp(self):
        self.settings_override = override_settings(MEDIA_ROOT=self.media_root)
        self.settings_override.enable()
        self.user = User.objects.create_user(
            username="coreuser",
            password="StrongPass123!",
            role="admin",
        )
        self.client.force_login(self.user)

    def tearDown(self):
        self.settings_override.disable()

    def test_dashboard_renders(self):
        response = self.client.get(reverse("dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Recent Lab Requests")

    def test_post_creates_patient_and_lab_request(self):
        response = self.client.post(
            reverse("request_create"),
            {
                "patient_full_name": "Juan Dela Cruz",
                "patient_sex": "male",
                "patient_birth_date": "1990-03-10",
                "age_snapshot_text": "",
                "request_datetime": "2026-03-10T09:30",
                "case_number": "CASE-001",
                "notes": "Walk-in patient",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(Patient.objects.count(), 1)
        self.assertEqual(LabRequest.objects.count(), 1)

        patient = Patient.objects.get()
        lab_request = LabRequest.objects.get()

        self.assertEqual(patient.full_name, "Juan Dela Cruz")
        self.assertEqual(lab_request.patient, patient)
        self.assertEqual(lab_request.patient_name_snapshot, "Juan Dela Cruz")
        self.assertEqual(lab_request.case_number, "CASE-001")
        self.assertTrue(lab_request.request_no.startswith("REQ-"))

    def test_request_form_defaults_single_active_facility_and_captures_branding_snapshot(self):
        organization = Organization.objects.create(
            legal_name="Naic Doctors Hospital Inc.",
            display_name="Naic Doctors Hospital Inc.",
            active=True,
        )
        facility = Facility.objects.create(
            organization=organization,
            display_name="Naic Main Facility",
            address="Brgy. Makina Naic, Cavite",
            contact_numbers="(046) 412-1443 / (046) 507-1510",
            report_header_image=SimpleUploadedFile(
                "brand.png",
                b"fake-brand-image",
                content_type="image/png",
            ),
            active=True,
        )

        get_response = self.client.get(reverse("request_create"))
        self.assertEqual(get_response.status_code, 200)
        self.assertEqual(get_response.context["form"].initial["facility"], facility)

        post_response = self.client.post(
            reverse("request_create"),
            {
                "patient_full_name": "Ana Reyes",
                "patient_sex": "female",
                "patient_birth_date": "1992-05-20",
                "age_snapshot_text": "",
                "request_datetime": "2026-03-10T10:15",
                "facility": facility.pk,
                "case_number": "CASE-002",
                "notes": "",
            },
        )

        self.assertEqual(post_response.status_code, 302)
        lab_request = LabRequest.objects.get(case_number="CASE-002")
        self.assertEqual(lab_request.facility, facility)
        self.assertEqual(lab_request.organization_name_snapshot, "Naic Doctors Hospital Inc.")
        self.assertEqual(lab_request.facility_name_snapshot, "Naic Main Facility")
        self.assertEqual(lab_request.facility_address_snapshot, "Brgy. Makina Naic, Cavite")
        self.assertEqual(lab_request.facility_contact_numbers_snapshot, "(046) 412-1443 / (046) 507-1510")
        self.assertTrue(bool(lab_request.facility_header_image_snapshot))

    def test_exam_options_endpoint_returns_latest_published_options(self):
        from apps.common.choices import ExamVersionStatusChoices
        from apps.exams.models import ExamDefinition, ExamDefinitionVersion, ExamOption

        exam_definition = ExamDefinition.objects.create(
            exam_code="chemistry",
            exam_name="Chemistry",
            active=True,
        )
        version = ExamDefinitionVersion.objects.create(
            exam_definition=exam_definition,
            version_no=1,
            version_status=ExamVersionStatusChoices.PUBLISHED,
            source_type="test",
            source_reference="test-v1",
        )
        ExamOption.objects.create(
            exam_version=version,
            option_key="fasting",
            option_label="Fasting",
            sort_order=1,
            active=True,
        )

        response = self.client.get(reverse("exam_definition_options", args=[exam_definition.pk]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["requires_option"])
        self.assertEqual(payload["options"][0]["label"], "Fasting")

    def test_admin_can_create_physician_from_custom_admin_portal(self):
        response = self.client.post(
            reverse("physician_create"),
            {
                "physician_code": "PHY-001",
                "display_name": "Dr. Sample",
                "active": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("physician_list"))

    def test_physician_list_filters_by_search_and_status(self):
        Physician.objects.create(physician_code="PHY-001", display_name="Dr. Active", active=True)
        Physician.objects.create(physician_code="PHY-002", display_name="Dr. Hidden", active=False)

        response = self.client.get(
            reverse("physician_list"),
            {
                "q": "Active",
                "status": "active",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Dr. Active")
        self.assertNotContains(response, "Dr. Hidden")


class MasterDataImportTests(TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.temp_dir = tempfile.mkdtemp()

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.temp_dir, ignore_errors=True)
        super().tearDownClass()

    def setUp(self):
        self.user = User.objects.create_user(
            username="importadmin",
            password="StrongPass123!",
            role="admin",
        )
        self.client.force_login(self.user)

    def build_master_data_workbook(self):
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Sheet A"
        ws["A1"] = "Field"
        ws["B1"] = "Input Type"
        ws["C1"] = "Dropdown List (Options)"
        ws["A8"] = "Requesting Physician"
        ws["C8"] = "DR. FIRST\nDR. SECOND\nDR. THIRD"
        ws["A9"] = "Room"
        ws["C9"] = "ER\nOPD \nNICU"
        ws["A14"] = "Medical Technologist"
        ws["C14"] = (
            "Alpha Medtech, RMT\nLic. No: 001\n\n"
            "Beta Medtech, RMT\nLic. No: 002"
        )
        ws["A16"] = "Pathologist"
        ws["C16"] = "Patho One, MD\nLIC. NO. 777"

        ws2 = workbook.create_sheet("Sheet B")
        ws2["A1"] = "Field"
        ws2["B1"] = "Input Type"
        ws2["C1"] = "Dropdown List (Options)"
        ws2["A8"] = "Requesting Physician"
        ws2["C8"] = "DR. SECOND\nDR. FOURTH"
        ws2["A9"] = "Room"
        ws2["C9"] = "ER\nWARD 1"
        ws2["A14"] = "Medical Technologist"
        ws2["C14"] = (
            "Alpha Medtech, RMT\nLic. No: 001\n\n"
            "Gamma Medtech, RMT\nLic. No: 003"
        )

        workbook_path = Path(self.temp_dir) / "master-data.xlsx"
        workbook.save(workbook_path)
        return workbook_path

    def test_import_master_data_creates_and_reactivates_records(self):
        workbook_path = self.build_master_data_workbook()

        Physician.objects.create(display_name="DR. SECOND", active=False)
        Room.objects.create(display_name="WARD 1", active=False)
        Signatory.objects.create(
            signatory_type=SignatoryTypeChoices.MEDTECH,
            display_name="Gamma Medtech, RMT",
            license_no="",
            active=False,
        )

        stats = import_master_data(workbook_path)

        self.assertEqual(stats.physician_source_count, 4)
        self.assertEqual(stats.room_source_count, 4)
        self.assertEqual(stats.signatory_source_count, 4)
        self.assertEqual(stats.physicians_created, 3)
        self.assertEqual(stats.physicians_reactivated, 1)
        self.assertEqual(stats.rooms_created, 3)
        self.assertEqual(stats.rooms_reactivated, 1)
        self.assertEqual(stats.signatories_created, 3)
        self.assertEqual(stats.signatories_reactivated, 1)
        self.assertEqual(stats.signatories_license_filled, 1)
        self.assertEqual(Physician.objects.count(), 4)
        self.assertEqual(Room.objects.count(), 4)
        self.assertEqual(Signatory.objects.count(), 4)
        self.assertTrue(Physician.objects.get(display_name="DR. SECOND").active)
        self.assertTrue(Room.objects.get(display_name="WARD 1").active)
        self.assertEqual(
            Signatory.objects.get(display_name="Gamma Medtech, RMT").license_no,
            "003",
        )
        self.assertTrue(Room.objects.filter(display_name="OPD").exists())

    def test_import_master_data_keeps_existing_conflicting_license_and_warns(self):
        workbook_path = self.build_master_data_workbook()
        Signatory.objects.create(
            signatory_type=SignatoryTypeChoices.PATHOLOGIST,
            display_name="Patho One, MD",
            license_no="OLD-LIC",
            active=True,
        )

        stats = import_master_data(workbook_path)

        self.assertTrue(stats.warnings)
        self.assertEqual(
            Signatory.objects.get(display_name="Patho One, MD").license_no,
            "OLD-LIC",
        )

    def test_admin_portal_import_page_runs_import_and_shows_summary(self):
        workbook_path = self.build_master_data_workbook()

        response = self.client.post(
            reverse("master_data_import"),
            {"workbook_path": str(workbook_path)},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Master data import finished.")
        self.assertContains(response, "Last Import Summary")
        self.assertContains(response, "Review Physicians")
        self.assertEqual(Physician.objects.count(), 4)
        self.assertEqual(Room.objects.count(), 4)
        self.assertEqual(Signatory.objects.count(), 4)

    def test_import_master_data_command_outputs_summary(self):
        workbook_path = self.build_master_data_workbook()
        stdout = StringIO()

        call_command(
            "import_master_data_workbook",
            file=str(workbook_path),
            stdout=stdout,
        )

        output = stdout.getvalue()
        self.assertIn("Master data import completed.", output)
        self.assertIn("Physicians found: 4", output)
        self.assertIn("Rooms found: 4", output)
        self.assertIn("Signatories found: 4", output)
