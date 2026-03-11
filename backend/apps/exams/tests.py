import tempfile
from pathlib import Path

from django.contrib.auth import get_user_model
from django.test import SimpleTestCase, TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook

from apps.common.choices import (
    ExamFieldDataTypeChoices,
    ExamFieldInputTypeChoices,
    ExamRuleTypeChoices,
    ExamVersionStatusChoices,
    RenderLayoutTypeChoices,
    UserRoleChoices,
)
from apps.exams.builder import create_draft_version, publish_draft_version, validate_draft_version
from apps.exams.models import ExamDefinition, ExamDefinitionVersion, ExamField, ExamFieldSelectOption, ExamOption, ExamRenderProfile, ExamRule, ExamSection
from apps.exams.services.workbook_import import (
    default_render_profile,
    infer_field_types,
    import_workbook,
    is_internal_note_row,
    make_field_key,
    parse_reference_range,
    sheet_payload,
)

User = get_user_model()


class WorkbookImportHelpersTests(SimpleTestCase):
    def test_make_field_key_prefixes_section(self):
        self.assertEqual(
            make_field_key("typhidot", "IgM"),
            "typhidot_igm",
        )

    def test_parse_between_reference_range(self):
        parsed = parse_reference_range("3.5 - 5.3 mEq/L")
        self.assertIsNotNone(parsed)
        self.assertEqual(parsed["range_type"], "numeric_between")
        self.assertEqual(str(parsed["min_numeric"]), "3.5")
        self.assertEqual(str(parsed["max_numeric"]), "5.3")

    def test_parse_less_than_reference_range(self):
        parsed = parse_reference_range("< 200 mg/dl")
        self.assertIsNotNone(parsed)
        self.assertEqual(parsed["range_type"], "numeric_less_than")
        self.assertEqual(str(parsed["max_numeric"]), "200")

    def test_infer_grouped_measurement_field(self):
        input_type, data_type = infer_field_types(
            "BBANK - Blood Bank",
            "VITAL SIGNS",
            "Predefined Selection",
            "BLOOD PRESSURE: mmHg",
            "",
        )
        self.assertEqual(input_type, ExamFieldInputTypeChoices.GROUPED_MEASUREMENT)
        self.assertEqual(data_type, ExamFieldDataTypeChoices.JSON)

    def test_note_rows_are_treated_as_internal_only(self):
        self.assertTrue(is_internal_note_row("NOTE"))

    def test_default_render_profile_sets_abg_variant(self):
        layout_type, config = default_render_profile(
            "ABG - Blood Gas Analysis",
            has_sections=True,
            has_reference_ranges=True,
        )

        self.assertEqual(layout_type, RenderLayoutTypeChoices.SECTIONED_REPORT)
        self.assertEqual(config["render_variant"], "abg_compact")
        self.assertEqual(config["left_section_key"], "blood_gas_value_abg")

    def test_default_render_profile_sets_bbank_variant(self):
        layout_type, config = default_render_profile(
            "BBANK - Blood Bank",
            has_sections=True,
            has_reference_ranges=True,
        )

        self.assertEqual(layout_type, RenderLayoutTypeChoices.SECTIONED_REPORT)
        self.assertEqual(config["render_variant"], "bbank_crossmatch")
        self.assertFalse(config["show_reference_ranges"])

    def test_default_render_profile_sets_serology_variant(self):
        layout_type, config = default_render_profile(
            "SEROLOGY",
            has_sections=True,
            has_reference_ranges=False,
        )

        self.assertEqual(layout_type, RenderLayoutTypeChoices.SECTIONED_REPORT)
        self.assertEqual(config["render_variant"], "serology_panel")
        self.assertEqual(config["option_to_section_keys"]["dengue_test"], "dengue_test")
        self.assertEqual(config["option_to_field_keys"]["hbsag_screening"], ["hbsag_screening"])

    def test_default_render_profile_sets_ogtt_variant(self):
        layout_type, config = default_render_profile(
            "OGTT - Blood Chemistry",
            has_sections=True,
            has_reference_ranges=True,
        )

        self.assertEqual(layout_type, RenderLayoutTypeChoices.SECTIONED_REPORT)
        self.assertEqual(config["render_variant"], "ogtt_timeline")
        self.assertEqual(config["option_to_section_keys"]["75g_ogtt"], "75g_oral_glucose_tolerance")
        self.assertEqual(config["option_to_field_keys"]["2_hour_postprandial"], ["2_hours_post_prandial"])

    def test_default_render_profile_sets_cardiaci_variant(self):
        layout_type, config = default_render_profile(
            "CARDIACI - Serology",
            has_sections=False,
            has_reference_ranges=True,
        )

        self.assertEqual(layout_type, RenderLayoutTypeChoices.RESULT_TABLE)
        self.assertEqual(config["render_variant"], "serology_panel")
        self.assertEqual(config["option_to_field_keys"]["ck_mb_tni_bnp"], ["ck_mb", "troponin_i", "bnp"])

    def test_default_render_profile_sets_hematology_variant(self):
        layout_type, config = default_render_profile(
            "HEMATOLOGY",
            has_sections=False,
            has_reference_ranges=True,
        )

        self.assertEqual(layout_type, RenderLayoutTypeChoices.RESULT_TABLE)
        self.assertEqual(config["render_variant"], "hematology_panel")
        self.assertEqual(config["sex_specific_field_map"]["hemoglobin"]["female"], "hemoglobin_f")
        self.assertEqual(
            config["option_to_panels"]["cbc_platelet_count_esr"][2]["keys"],
            ["esr"],
        )

    def test_default_render_profile_sets_bcmale_variant(self):
        layout_type, config = default_render_profile(
            "BCMALE - Blood Chemistry",
            has_sections=False,
            has_reference_ranges=True,
        )

        self.assertEqual(layout_type, RenderLayoutTypeChoices.RESULT_TABLE)
        self.assertEqual(config["render_variant"], "chemistry_panel")
        self.assertEqual(config["panel_groups"][0]["keys"], ["fasting_blood_sugar", "random_blood_sugar", "hgt"])
        self.assertEqual(config["panel_groups"][3]["title"], "Lipid Profile")

    def test_default_render_profile_sets_bcfemale_variant(self):
        layout_type, config = default_render_profile(
            "BCFEMALE - Blood Chemistry",
            has_sections=False,
            has_reference_ranges=True,
        )

        self.assertEqual(layout_type, RenderLayoutTypeChoices.RESULT_TABLE)
        self.assertEqual(config["render_variant"], "chemistry_panel")
        self.assertEqual(config["panel_groups"][1]["keys"], ["blood_urea_nitrogen", "creatinine", "blood_uric_acid"])

    def test_default_render_profile_sets_urine_variant(self):
        layout_type, config = default_render_profile(
            "URINE- Clinical Microscopy",
            has_sections=True,
            has_reference_ranges=False,
        )

        self.assertEqual(layout_type, RenderLayoutTypeChoices.SECTIONED_REPORT)
        self.assertEqual(config["render_variant"], "microscopy_sections")
        self.assertEqual(
            config["option_to_sections"]["urinalysis"],
            ["macroscopic_finding", "clinical_finding", "microscopic_finding"],
        )
        self.assertEqual(
            config["option_to_field_keys"]["pregnancy_test"],
            ["microscopic_finding_pregnancy_test"],
        )

    def test_default_render_profile_sets_fecalysis_variant(self):
        layout_type, config = default_render_profile(
            "FECALYSIS - Clinical Microscopy",
            has_sections=True,
            has_reference_ranges=False,
        )

        self.assertEqual(layout_type, RenderLayoutTypeChoices.SECTIONED_REPORT)
        self.assertEqual(config["render_variant"], "microscopy_sections")
        self.assertEqual(
            config["option_to_sections"]["fecalysis"],
            ["macroscopic_finding", "microscopic_finding", "microscopic_finding_2"],
        )
        self.assertEqual(
            config["option_to_field_keys"]["fobt"],
            ["macroscopic_finding_fecal_occult_blood"],
        )

    def test_default_render_profile_sets_protime_variant(self):
        layout_type, config = default_render_profile(
            "PROTIME, APTT - Hematology",
            has_sections=True,
            has_reference_ranges=True,
        )

        self.assertEqual(layout_type, RenderLayoutTypeChoices.SECTIONED_REPORT)
        self.assertEqual(config["render_variant"], "coagulation_panel")
        self.assertEqual(config["option_to_sections"]["protime_aptt"], ["pro_time", "aptt"])

    def test_default_render_profile_sets_semen_variant(self):
        layout_type, config = default_render_profile(
            "SEMEN - Clinical Microscopy",
            has_sections=True,
            has_reference_ranges=False,
        )

        self.assertEqual(layout_type, RenderLayoutTypeChoices.SECTIONED_REPORT)
        self.assertEqual(config["render_variant"], "semen_analysis")
        self.assertEqual(
            config["sample_field_keys"],
            ["time_collected", "time_received", "total_volume", "liquefaction_time"],
        )
        self.assertEqual(config["section_keys"], ["motility", "morphology", "sperm_count", "others"])

    def test_default_render_profile_sets_microbiology_variant(self):
        layout_type, config = default_render_profile(
            "MICROBIOLOGY",
            has_sections=False,
            has_reference_ranges=False,
        )

        self.assertEqual(layout_type, RenderLayoutTypeChoices.LABEL_VALUE_LIST)
        self.assertEqual(config["render_variant"], "single_result_focus")
        self.assertEqual(config["field_keys"], ["result"])

    def test_default_render_profile_sets_hba1c_variant(self):
        layout_type, config = default_render_profile(
            "HBA1C - Blood Chemistry",
            has_sections=False,
            has_reference_ranges=True,
        )

        self.assertEqual(layout_type, RenderLayoutTypeChoices.RESULT_TABLE)
        self.assertEqual(config["render_variant"], "single_result_focus")
        self.assertTrue(config["show_reference_ranges"])
        self.assertEqual(config["field_keys"], ["result"])

    def test_default_render_profile_sets_hiv_variant(self):
        layout_type, config = default_render_profile(
            "HIV 1&2 TESTING - Serology",
            has_sections=False,
            has_reference_ranges=False,
        )

        self.assertEqual(layout_type, RenderLayoutTypeChoices.LABEL_VALUE_LIST)
        self.assertEqual(config["render_variant"], "rapid_test_panel")
        self.assertEqual(config["meta_field_keys"], ["lot_number"])
        self.assertEqual(config["result_field_keys"], ["test_result"])

    def test_default_render_profile_sets_covid_variant(self):
        layout_type, config = default_render_profile(
            "COVID 19 ANTIGEN (RAPID TEST) -",
            has_sections=False,
            has_reference_ranges=False,
        )

        self.assertEqual(layout_type, RenderLayoutTypeChoices.LABEL_VALUE_LIST)
        self.assertEqual(config["render_variant"], "rapid_test_panel")
        self.assertEqual(config["result_field_keys"], ["test_result"])
        self.assertEqual(config["attachment_field_keys"], ["result_image"])


class WorkbookImportIntegrationTests(TestCase):
    def write_workbook(self, path, sheet_title, rows, exam_options=None, headers=None):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_title
        headers = headers or ["Field", "Input Type", "Selection / Unit", "Normal Range", "Notes"]
        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(1, col_idx, header)

        if exam_options:
            worksheet.cell(7, 3, "\n".join(exam_options))

        for row_index, row in enumerate(rows, start=2):
            field, input_type, raw_options, reference, notes = row
            worksheet.cell(row_index, 1, field)
            worksheet.cell(row_index, 2, input_type)
            worksheet.cell(row_index, 3, raw_options)
            worksheet.cell(row_index, 4, reference)
            worksheet.cell(row_index, 5, notes)

        workbook.save(path)

    def test_serology_hbsag_is_not_scoped_to_typhidot_section(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "serology.xlsx"
            self.write_workbook(
                workbook_path,
                "SEROLOGY",
                [
                    ("TYPHIDOT", "", "", "", ""),
                    ("IgM", "Manual Entry", "", "", ""),
                    ("HbsAg SCREENING:", "Manual Entry", "", "", ""),
                ],
            )

            import_workbook(workbook_path)

        exam = ExamDefinition.objects.get(exam_code="serology")
        version = exam.versions.get(version_no=1)
        igm_field = version.fields.get(field_key="typhidot_igm")
        hbsag_field = version.fields.get(field_key="hbsag_screening")

        self.assertEqual(igm_field.section.section_key, "typhidot")
        self.assertIsNone(hbsag_field.section)

    def test_ogtt_post_prandial_is_not_scoped_to_100g_section(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "ogtt.xlsx"
            self.write_workbook(
                workbook_path,
                "OGTT - Blood Chemistry",
                [
                    ("100G ORAL GLUCOSE TOLERANCE", "", "", "", ""),
                    ("1ST HOUR", "Manual Entry", "", "70 - 105", ""),
                    ("2 HOURS POST PRANDIAL", "Manual Entry", "", "70 - 140", ""),
                ],
            )

            import_workbook(workbook_path)

        exam = ExamDefinition.objects.get(exam_code="ogtt")
        version = exam.versions.get(version_no=1)
        first_hour_field = version.fields.get(field_key="100g_oral_glucose_tolerance_1st_hour")
        post_prandial_field = version.fields.get(field_key="2_hours_post_prandial")

        self.assertEqual(first_hour_field.section.section_key, "100g_oral_glucose_tolerance")
        self.assertIsNone(post_prandial_field.section)

    def test_force_reimport_creates_new_version_for_same_workbook(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "serology.xlsx"
            self.write_workbook(
                workbook_path,
                "SEROLOGY",
                [
                    ("HbsAg SCREENING:", "Manual Entry", "", "", ""),
                ],
            )

            first_stats = import_workbook(workbook_path)
            second_stats = import_workbook(workbook_path)
            forced_stats = import_workbook(workbook_path, force=True)

        exam = ExamDefinition.objects.get(exam_code="serology")

        self.assertEqual(first_stats.created_versions, 1)
        self.assertEqual(second_stats.skipped_versions, 1)
        self.assertEqual(forced_stats.created_versions, 1)
        self.assertEqual(exam.versions.count(), 2)
        self.assertEqual(exam.versions.order_by("-version_no").first().version_no, 2)

    def test_grouped_measurement_field_keeps_subfield_config(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "bbank.xlsx"
            self.write_workbook(
                workbook_path,
                "BBANK - Blood Bank",
                [
                    ("TYPE OF CROSSMATCHING", "", "", "", ""),
                    (
                        "VITAL SIGNS",
                        "Predefined Selection",
                        "BLOOD PRESSURE: mmHg\nPULSE RATE: bpm\n",
                        "Use dropdowns plus numeric entry",
                        "",
                    ),
                ],
            )

            import_workbook(workbook_path)

        exam = ExamDefinition.objects.get(exam_code="bbank")
        version = exam.versions.get(version_no=1)
        field = version.fields.get(field_key="type_of_crossmatching_vital_signs")

        self.assertEqual(field.input_type, ExamFieldInputTypeChoices.GROUPED_MEASUREMENT)
        self.assertEqual(
            field.config_json["grouped_fields"],
            [
                {
                    "key": "blood_pressure",
                    "label": "BLOOD PRESSURE",
                    "unit": "mmHg",
                    "input_type": ExamFieldInputTypeChoices.TEXT,
                    "sort_order": 1,
                },
                {
                    "key": "pulse_rate",
                    "label": "PULSE RATE",
                    "unit": "bpm",
                    "input_type": ExamFieldInputTypeChoices.TEXT,
                    "sort_order": 2,
                },
            ],
        )

    def test_note_rows_are_not_imported_and_notes_column_stays_internal(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "abg.xlsx"
            self.write_workbook(
                workbook_path,
                "ABG - Blood Gas Analysis",
                [
                    ("Blood gas value (ABG)", "", "", "", ""),
                    ("pH", "Manual Entry", "", "7.35 - 7.45", "Do not show to users"),
                    ("NOTE", "Manual Entry", "", "", "All abnormal values print red"),
                ],
            )

            import_workbook(workbook_path)

        exam = ExamDefinition.objects.get(exam_code="abg")
        version = exam.versions.get(version_no=1)
        field = version.fields.get(field_key="blood_gas_value_abg_ph")

        self.assertEqual(version.fields.count(), 1)
        self.assertEqual(field.help_text, "")
        self.assertEqual(field.config_json["internal_note"], "Do not show to users")

    def test_header_aware_notes_column_does_not_become_reference_text(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "bbank.xlsx"
            self.write_workbook(
                workbook_path,
                "BBANK - Blood Bank",
                [
                    ("TYPE OF CROSSMATCHING", "", "", "", ""),
                    (
                        "VITAL SIGNS",
                        "Predefined Selection",
                        "BLOOD PRESSURE: mmHg\nPULSE RATE: bpm\n",
                        "Use dropdowns plus numeric entry",
                        "",
                    ),
                ],
                headers=["Field", "Input Type", "Dropdown List (Options)", "Notes", None],
            )

            import_workbook(workbook_path)

        exam = ExamDefinition.objects.get(exam_code="bbank")
        version = exam.versions.get(version_no=1)
        field = version.fields.get(field_key="type_of_crossmatching_vital_signs")

        self.assertEqual(field.reference_text, "")
        self.assertEqual(field.config_json["internal_note"], "Use dropdowns plus numeric entry")

    def test_sheet_payload_ignores_meaningless_whitespace_rows(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "SEROLOGY"
        worksheet.cell(1, 1, "Field")
        worksheet.cell(1, 2, "Input Type")
        worksheet.cell(1, 3, "Dropdown List (Options)")
        worksheet.cell(1, 4, "Notes")
        worksheet.cell(2, 1, "HbsAg SCREENING:")
        worksheet.cell(2, 2, "Manual Entry")
        worksheet.cell(500, 1, "   ")

        payload = sheet_payload(worksheet)

        self.assertEqual(len(payload["rows"]), 2)
        self.assertEqual(payload["rows"][1]["field"], "HbsAg SCREENING:")

    def test_whitespace_only_rows_do_not_force_new_import_version(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "serology.xlsx"
            self.write_workbook(
                workbook_path,
                "SEROLOGY",
                [
                    ("HbsAg SCREENING:", "Manual Entry", "", "", ""),
                ],
                headers=["Field", "Input Type", "Dropdown List (Options)", "Notes", None],
            )

            first_stats = import_workbook(workbook_path)

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "SEROLOGY"
            worksheet.cell(1, 1, "Field")
            worksheet.cell(1, 2, "Input Type")
            worksheet.cell(1, 3, "Dropdown List (Options)")
            worksheet.cell(1, 4, "Notes")
            worksheet.cell(2, 1, "HbsAg SCREENING:")
            worksheet.cell(2, 2, "Manual Entry")
            worksheet.cell(500, 1, "   ")
            workbook.save(workbook_path)

            second_stats = import_workbook(workbook_path)

        exam = ExamDefinition.objects.get(exam_code="serology")

        self.assertEqual(first_stats.created_versions, 1)
        self.assertEqual(second_stats.skipped_versions, 1)
        self.assertEqual(exam.versions.count(), 1)


class ExamBuilderTests(TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_user(
            username="examadmin",
            password="StrongPass123!",
            role=UserRoleChoices.ADMIN,
        )
        self.client.force_login(self.admin_user)

    def test_exam_definition_create_view_builds_initial_draft_and_render_profile(self):
        response = self.client.post(
            reverse("exam_definition_create"),
            {
                "exam_code": "",
                "exam_name": "Custom Viral Test",
                "category": "Serology",
                "description": "Admin-created custom exam",
                "active": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        definition = ExamDefinition.objects.get(exam_name="Custom Viral Test")
        draft = definition.versions.get(version_status=ExamVersionStatusChoices.DRAFT)
        self.assertEqual(definition.exam_code, "custom-viral-test")
        self.assertEqual(draft.version_no, 1)
        self.assertEqual(draft.source_reference, "manual-builder:new-exam")
        self.assertTrue(hasattr(draft, "render_profile"))

    def test_create_draft_version_clones_options_sections_fields_ranges_rules_and_render_profile(self):
        definition = ExamDefinition.objects.create(exam_code="clone-demo", exam_name="Clone Demo", active=True)
        published_version = ExamDefinitionVersion.objects.create(
            exam_definition=definition,
            version_no=1,
            version_status=ExamVersionStatusChoices.PUBLISHED,
            source_type="test",
            source_reference="test-v1",
            published_at=timezone.now(),
        )
        option = ExamOption.objects.create(
            exam_version=published_version,
            option_key="panel_a",
            option_label="Panel A",
            sort_order=1,
            active=True,
        )
        section = ExamSection.objects.create(
            exam_version=published_version,
            section_key="main_results",
            section_label="Main Results",
            sort_order=1,
            active=True,
        )
        field = ExamField.objects.create(
            exam_version=published_version,
            section=section,
            field_key="result_value",
            field_label="Result Value",
            input_type=ExamFieldInputTypeChoices.SELECT,
            data_type=ExamFieldDataTypeChoices.STRING,
            sort_order=1,
            active=True,
        )
        ExamFieldSelectOption.objects.create(
            field=field,
            option_value="positive",
            option_label="Positive",
            sort_order=1,
            active=True,
        )
        field.reference_ranges.create(
            option_scope=option,
            range_type="text_only",
            reference_text="Non-reactive",
            sort_order=1,
        )
        ExamRule.objects.create(
            exam_version=published_version,
            rule_type=ExamRuleTypeChoices.VISIBILITY,
            target_type="field",
            target_id=field.id,
            condition_json={"exam_option_keys": ["panel_a"]},
            effect_json={"visible": True},
            sort_order=1,
            active=True,
        )
        ExamRenderProfile.objects.create(
            exam_version=published_version,
            layout_type=RenderLayoutTypeChoices.RESULT_TABLE,
            config_json={"show_units": False, "show_reference_ranges": True, "render_variant": "serology_panel"},
            active=True,
        )

        draft, created = create_draft_version(definition, user=self.admin_user)

        self.assertTrue(created)
        self.assertEqual(draft.version_status, ExamVersionStatusChoices.DRAFT)
        self.assertEqual(draft.version_no, 2)
        self.assertEqual(draft.options.count(), 1)
        self.assertEqual(draft.sections.count(), 1)
        self.assertEqual(draft.fields.count(), 1)
        cloned_field = draft.fields.get()
        self.assertEqual(cloned_field.section.section_key, "main_results")
        self.assertEqual(cloned_field.select_options.get().option_label, "Positive")
        self.assertEqual(cloned_field.reference_ranges.get().option_scope.option_key, "panel_a")
        self.assertEqual(draft.rules.get().target_id, cloned_field.id)
        self.assertEqual(draft.render_profile.config_json["render_variant"], "serology_panel")

    def test_publish_draft_version_archives_old_published_version(self):
        definition = ExamDefinition.objects.create(exam_code="publish-demo", exam_name="Publish Demo", active=True)
        published_version = ExamDefinitionVersion.objects.create(
            exam_definition=definition,
            version_no=1,
            version_status=ExamVersionStatusChoices.PUBLISHED,
            source_type="test",
            source_reference="test-v1",
            published_at=timezone.now(),
        )
        draft = ExamDefinitionVersion.objects.create(
            exam_definition=definition,
            version_no=2,
            version_status=ExamVersionStatusChoices.DRAFT,
            source_type="manual_builder",
            source_reference="manual-builder:new-exam",
        )
        field = ExamField.objects.create(
            exam_version=draft,
            field_key="result_value",
            field_label="Result Value",
            input_type=ExamFieldInputTypeChoices.TEXT,
            data_type=ExamFieldDataTypeChoices.STRING,
            sort_order=1,
            active=True,
        )
        ExamRenderProfile.objects.create(
            exam_version=draft,
            layout_type=RenderLayoutTypeChoices.RESULT_TABLE,
            config_json={"show_units": True, "show_reference_ranges": True},
            active=True,
        )

        publish_draft_version(draft, user=self.admin_user, change_summary="Released new custom layout.")

        published_version.refresh_from_db()
        draft.refresh_from_db()
        self.assertEqual(published_version.version_status, ExamVersionStatusChoices.ARCHIVED)
        self.assertEqual(draft.version_status, ExamVersionStatusChoices.PUBLISHED)
        self.assertEqual(draft.published_by, self.admin_user)
        self.assertEqual(draft.change_summary, "Released new custom layout.")

    def test_validate_draft_version_blocks_select_field_without_choices(self):
        definition = ExamDefinition.objects.create(exam_code="validation-demo", exam_name="Validation Demo", active=True)
        draft = ExamDefinitionVersion.objects.create(
            exam_definition=definition,
            version_no=1,
            version_status=ExamVersionStatusChoices.DRAFT,
            source_type="manual_builder",
            source_reference="manual-builder:new-exam",
        )
        ExamField.objects.create(
            exam_version=draft,
            field_key="qualitative_result",
            field_label="Qualitative Result",
            input_type=ExamFieldInputTypeChoices.SELECT,
            data_type=ExamFieldDataTypeChoices.STRING,
            sort_order=1,
            active=True,
        )
        ExamRenderProfile.objects.create(
            exam_version=draft,
            layout_type=RenderLayoutTypeChoices.RESULT_TABLE,
            config_json={"show_units": True, "show_reference_ranges": True},
            active=True,
        )

        errors = validate_draft_version(draft)

        self.assertIn("Select field 'Qualitative Result' needs at least one active choice.", errors)

    def test_exam_rule_create_view_serializes_conditions(self):
        definition = ExamDefinition.objects.create(exam_code="rule-demo", exam_name="Rule Demo", active=True)
        draft = ExamDefinitionVersion.objects.create(
            exam_definition=definition,
            version_no=1,
            version_status=ExamVersionStatusChoices.DRAFT,
            source_type="manual_builder",
            source_reference="manual-builder:new-exam",
        )
        option = ExamOption.objects.create(
            exam_version=draft,
            option_key="panel_a",
            option_label="Panel A",
            sort_order=1,
            active=True,
        )
        field = ExamField.objects.create(
            exam_version=draft,
            field_key="result_value",
            field_label="Result Value",
            input_type=ExamFieldInputTypeChoices.TEXT,
            data_type=ExamFieldDataTypeChoices.STRING,
            sort_order=1,
            active=True,
        )

        response = self.client.post(
            reverse("exam_rule_create", args=[draft.pk]),
            {
                "rule_type": ExamRuleTypeChoices.REQUIREMENT,
                "target_type": "field",
                "target_field": field.pk,
                "target_section": "",
                "option_scopes": [option.pk],
                "sex_scope": "female",
                "sort_order": 1,
                "active": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        rule = draft.rules.get()
        self.assertEqual(rule.target_type, "field")
        self.assertEqual(rule.target_id, field.id)
        self.assertEqual(rule.condition_json, {"exam_option_keys": ["panel_a"], "patient_sex": ["female"]})
        self.assertEqual(rule.effect_json, {"required": True})
