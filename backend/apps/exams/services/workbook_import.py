import hashlib
import json
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import load_workbook

from apps.common.choices import (
    ExamFieldDataTypeChoices,
    ExamFieldInputTypeChoices,
    ExamRuleTypeChoices,
    ExamVersionStatusChoices,
    RenderLayoutTypeChoices,
)
from apps.exams.models import (
    ExamDefinition,
    ExamDefinitionVersion,
    ExamField,
    ExamFieldReferenceRange,
    ExamFieldSelectOption,
    ExamOption,
    ExamRenderProfile,
    ExamRule,
    ExamSection,
)


PATIENT_FIELDS = {
    "Name",
    "Age",
    "Sex",
    "Date",
    "Date/Time",
    "Examination",
    "Requesting Physician",
    "Room",
    "Case Number",
}

SIGNATORY_FIELDS = {"Medical Technologist", "Pathologist"}
SKIP_SECTION_LABELS = {"PATIENT INFORMATION"}

TEXT_MANUAL_FIELDS = {
    "LOT NUMBER",
    "PATIENT’S BLOOD TYPE",
    "PATIENT'S BLOOD TYPE",
    "BLOOD COMPONENT",
    "DONOR’S BLOOD TYPE",
    "DONOR'S BLOOD TYPE",
    "SOURCE OF BLOOD",
    "SERIAL NUMBER",
    "REMARKS",
    "RELEASED BY",
    "RELEASED TO",
    "OTHERS",
    "OTHERS  ",
}

DATE_FIELDS = {
    "DATE EXTRACTED",
    "DATE EXPIRY",
}

SHEET_META_MAP = {
    "URINE- Clinical Microscopy": {"exam_code": "urine", "exam_name": "Urine", "category": "Clinical Microscopy"},
    "SEROLOGY": {"exam_code": "serology", "exam_name": "Serology", "category": "Serology"},
    "SEMEN - Clinical Microscopy": {"exam_code": "semen", "exam_name": "Semen Analysis", "category": "Clinical Microscopy"},
    "OGTT - Blood Chemistry": {"exam_code": "ogtt", "exam_name": "Oral Glucose Tolerance Test", "category": "Blood Chemistry"},
    "MICROBIOLOGY": {"exam_code": "microbiology", "exam_name": "Microbiology", "category": "Microbiology"},
    "HEMATOLOGY": {"exam_code": "hematology", "exam_name": "Hematology", "category": "Hematology"},
    "HBA1C - Blood Chemistry": {"exam_code": "hba1c", "exam_name": "HBA1C", "category": "Blood Chemistry"},
    "FECALYSIS - Clinical Microscopy": {"exam_code": "fecalysis", "exam_name": "Fecalysis", "category": "Clinical Microscopy"},
    "CARDIACI - Serology": {"exam_code": "cardiaci", "exam_name": "Cardiac Markers", "category": "Serology"},
    "BCMALE - Blood Chemistry": {"exam_code": "bcmale", "exam_name": "Blood Chemistry Male", "category": "Blood Chemistry"},
    "BCFEMALE - Blood Chemistry": {"exam_code": "bcfemale", "exam_name": "Blood Chemistry Female", "category": "Blood Chemistry"},
    "BBANK - Blood Bank": {"exam_code": "bbank", "exam_name": "Blood Bank", "category": "Blood Bank"},
    "ABG - Blood Gas Analysis": {"exam_code": "abg", "exam_name": "Blood Gas Analysis", "category": "Blood Gas Analysis"},
    "HIV 1&2 TESTING - Serology": {"exam_code": "hiv-1-2-testing", "exam_name": "HIV 1 and 2 Testing", "category": "Serology"},
    "COVID 19 ANTIGEN (RAPID TEST) -": {"exam_code": "covid-19-antigen-rapid-test", "exam_name": "COVID 19 Antigen Rapid Test", "category": "Serology"},
    "PROTIME, APTT - Hematology": {"exam_code": "protime-aptt", "exam_name": "Protime and APTT", "category": "Hematology"},
}

UNSCOPED_FIELDS_BY_SHEET = {
    "SEROLOGY": {
        "HbsAg SCREENING:",
        "VDRL:",
        "ANTI-HCV:",
        "ASO TITER:",
        "OTHERS",
    },
    "OGTT - Blood Chemistry": {
        "2 HOURS POST PRANDIAL",
        "50 G ORAL GLUCOSE CHALLENGE",
        "Others",
    },
}

IMPORTER_SIGNATURE_VERSION = 11


@dataclass
class ImportStats:
    created_definitions: int = 0
    created_versions: int = 0
    skipped_versions: int = 0
    created_options: int = 0
    created_sections: int = 0
    created_fields: int = 0
    created_ranges: int = 0
    created_rules: int = 0


@dataclass
class SheetColumnConfig:
    field_col: int = 1
    input_type_col: int = 2
    raw_options_col: int = 3
    reference_col: int | None = None
    notes_col: int | None = None
    sheet_notes: list[str] | None = None


HEADER_ALIASES = {
    "field": {"field"},
    "input_type": {"input type"},
    "raw_options": {"dropdown list (options)", "selection / unit"},
    "reference": {"normal value", "normal values", "normal range", "reference", "reference range"},
    "notes": {"notes", "note"},
}


def normalize_text(value):
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split()).strip()


def is_internal_note_row(field_label):
    return normalize_text(field_label).upper() == "NOTE"


def split_multiline(value):
    if not isinstance(value, str):
        return []
    parts = [normalize_text(part) for part in value.splitlines()]
    cleaned = [part for part in parts if part and set(part) != {"-"}]
    return cleaned


def normalize_header(value):
    return normalize_text(value).lower().rstrip(":")


def resolve_sheet_columns(ws):
    config = SheetColumnConfig(sheet_notes=[])

    for col_idx in range(1, min(ws.max_column, 5) + 1):
        header_value = ws.cell(1, col_idx).value
        normalized_header = normalize_header(header_value)
        if not normalized_header:
            continue

        matched_role = None
        for role, aliases in HEADER_ALIASES.items():
            if normalized_header in aliases:
                matched_role = role
                break

        if matched_role == "field":
            config.field_col = col_idx
        elif matched_role == "input_type":
            config.input_type_col = col_idx
        elif matched_role == "raw_options":
            config.raw_options_col = col_idx
        elif matched_role == "reference":
            config.reference_col = col_idx
        elif matched_role == "notes":
            config.notes_col = col_idx
        elif col_idx >= 4:
            config.sheet_notes.append(normalize_text(header_value))

    return config


def get_cell_text(ws, row_idx, col_idx):
    if not col_idx:
        return ""
    return normalize_text(ws.cell(row_idx, col_idx).value)


def get_cell_value(ws, row_idx, col_idx):
    if not col_idx:
        return ""
    return ws.cell(row_idx, col_idx).value or ""


def row_has_meaningful_content(field, input_type, raw_options, reference, notes):
    return any(
        [
            field,
            input_type,
            normalize_text(raw_options),
            reference,
            notes,
        ]
    )


def make_meta(sheet_name):
    if sheet_name in SHEET_META_MAP:
        return SHEET_META_MAP[sheet_name]

    cleaned = sheet_name.strip().rstrip("-").strip()
    return {
        "exam_code": slugify(cleaned),
        "exam_name": cleaned,
        "category": "",
    }


def make_key(text):
    key = slugify(normalize_text(text)).replace("-", "_")
    return key or "field"


def make_field_key(section_key, field_label):
    base_key = make_key(field_label)
    if section_key:
        return f"{section_key}_{base_key}"
    return base_key


def decimal_or_none(value):
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def parse_reference_range(text):
    raw = normalize_text(text)
    if not raw:
        return None

    between_match = re.search(r"([+-]?\d+(?:\.\d+)?)\s*(?:-|to)\s*([+-]?\d+(?:\.\d+)?)", raw, flags=re.IGNORECASE)
    if between_match:
        low = decimal_or_none(between_match.group(1))
        high = decimal_or_none(between_match.group(2))
        if low is not None and high is not None:
            return {
                "range_type": "numeric_between",
                "min_numeric": low,
                "max_numeric": high,
                "reference_text": raw,
                "abnormal_rule": "outside_range",
            }

    lt_match = re.search(r"(?:<|less than)\s*([+-]?\d+(?:\.\d+)?)", raw, flags=re.IGNORECASE)
    if lt_match:
        high = decimal_or_none(lt_match.group(1))
        if high is not None:
            return {
                "range_type": "numeric_less_than",
                "min_numeric": None,
                "max_numeric": high,
                "reference_text": raw,
                "abnormal_rule": "greater_than_max",
            }

    gt_match = re.search(r"(?:>|greater than)\s*([+-]?\d+(?:\.\d+)?)", raw, flags=re.IGNORECASE)
    if gt_match:
        low = decimal_or_none(gt_match.group(1))
        if low is not None:
            return {
                "range_type": "numeric_greater_than",
                "min_numeric": low,
                "max_numeric": None,
                "reference_text": raw,
                "abnormal_rule": "less_than_min",
            }

    return {
        "range_type": "text_reference",
        "min_numeric": None,
        "max_numeric": None,
        "reference_text": raw,
        "abnormal_rule": "",
    }


def infer_field_types(sheet_name, field_label, input_type, raw_options, reference_text):
    normalized_label = normalize_text(field_label).upper()
    normalized_input = normalize_text(input_type).lower()
    normalized_reference = normalize_text(reference_text)
    normalized_options = normalize_text(raw_options)

    if sheet_name == "BBANK - Blood Bank" and normalized_label == "VITAL SIGNS":
        return ExamFieldInputTypeChoices.GROUPED_MEASUREMENT, ExamFieldDataTypeChoices.JSON

    if normalized_input == "predefined selection":
        return ExamFieldInputTypeChoices.SELECT, ExamFieldDataTypeChoices.STRING

    if normalized_label in DATE_FIELDS:
        return ExamFieldInputTypeChoices.DATE, ExamFieldDataTypeChoices.DATE

    if "DATE/TIME" in normalized_label or normalized_label == "DATE TIME":
        return ExamFieldInputTypeChoices.DATETIME, ExamFieldDataTypeChoices.DATETIME

    if "DATE" in normalized_label and "TIME" not in normalized_label:
        return ExamFieldInputTypeChoices.DATE, ExamFieldDataTypeChoices.DATE

    if normalized_label in TEXT_MANUAL_FIELDS:
        return ExamFieldInputTypeChoices.TEXT, ExamFieldDataTypeChoices.STRING

    if normalized_label in {"OTHERS", "OTHERS  "}:
        return ExamFieldInputTypeChoices.TEXTAREA, ExamFieldDataTypeChoices.STRING

    if normalized_reference and re.search(r"\d", normalized_reference):
        return ExamFieldInputTypeChoices.DECIMAL, ExamFieldDataTypeChoices.DECIMAL

    if normalized_options and "\n" not in str(raw_options) and normalized_options not in {"------------------------------------", "-----------------------------"}:
        if re.search(r"[A-Za-z/%]", normalized_options):
            return ExamFieldInputTypeChoices.DECIMAL, ExamFieldDataTypeChoices.DECIMAL

    return ExamFieldInputTypeChoices.TEXT, ExamFieldDataTypeChoices.STRING


def extract_unit(input_type, raw_options):
    normalized_input = normalize_text(input_type).lower()
    if normalized_input != "manual entry":
        return ""

    value = normalize_text(raw_options)
    if not value or set(value) == {"-"}:
        return ""
    if "\n" in str(raw_options):
        return ""
    return value


def parse_grouped_measurement_config(raw_options):
    grouped_fields = []
    for position, line in enumerate(split_multiline(raw_options), start=1):
        label = line
        unit = ""
        if ":" in line:
            label_part, unit_part = line.split(":", 1)
            label = normalize_text(label_part)
            unit = normalize_text(unit_part)

        grouped_fields.append(
            {
                "key": make_key(label),
                "label": label,
                "unit": unit,
                "input_type": ExamFieldInputTypeChoices.TEXT,
                "sort_order": position,
            }
        )

    return grouped_fields


def build_field_config(field_input_type, raw_options, internal_note=""):
    raw_option_lines = split_multiline(raw_options)
    config = {}
    if raw_option_lines:
        config["raw_options_lines"] = raw_option_lines
    if internal_note:
        config["internal_note"] = internal_note

    if field_input_type == ExamFieldInputTypeChoices.GROUPED_MEASUREMENT:
        config["grouped_fields"] = parse_grouped_measurement_config(raw_options)

    return config


def sheet_payload(ws):
    column_config = resolve_sheet_columns(ws)
    rows = []
    for row_idx in range(1, ws.max_row + 1):
        field = get_cell_text(ws, row_idx, column_config.field_col)
        input_type = get_cell_text(ws, row_idx, column_config.input_type_col)
        raw_options = get_cell_value(ws, row_idx, column_config.raw_options_col)
        reference = get_cell_text(ws, row_idx, column_config.reference_col)
        notes = get_cell_text(ws, row_idx, column_config.notes_col)

        if not row_has_meaningful_content(field, input_type, raw_options, reference, notes):
            continue

        row = {
            "row": row_idx,
            "field": field,
            "input_type": input_type,
            "raw_options": raw_options,
            "reference": reference,
            "notes": notes,
        }
        rows.append(row)
    return {
        "rows": rows,
        "columns": {
            "field": column_config.field_col,
            "input_type": column_config.input_type_col,
            "raw_options": column_config.raw_options_col,
            "reference": column_config.reference_col,
            "notes": column_config.notes_col,
        },
        "sheet_notes": column_config.sheet_notes or [],
    }


def payload_hash(payload):
    encoded = json.dumps(payload, ensure_ascii=True, sort_keys=True, default=str).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def build_source_reference(sheet_name, payload, importer_signature_version=IMPORTER_SIGNATURE_VERSION):
    signature = {
        "sheet_name": sheet_name,
        "importer_signature_version": importer_signature_version,
        "payload": payload,
    }
    return f"v{importer_signature_version}:{payload_hash(signature)}"


def create_rule(version, rule_type, target_type, target_id, condition, effect, sort_order):
    return ExamRule.objects.create(
        exam_version=version,
        rule_type=rule_type,
        target_type=target_type,
        target_id=target_id,
        condition_json=condition,
        effect_json=effect,
        sort_order=sort_order,
        active=True,
    )


def build_safe_rules(sheet_name, version, options_by_key, sections_by_key, fields_by_key):
    created = 0
    sort_order = 0

    for field in fields_by_key.values():
        label = normalize_text(field.field_label).upper()
        if label.endswith("(M)"):
            sort_order += 1
            create_rule(
                version,
                ExamRuleTypeChoices.VISIBILITY,
                "field",
                field.id,
                {"patient_sex": ["male"]},
                {"visible": True},
                sort_order,
            )
            created += 1
        elif label.endswith("(F)"):
            sort_order += 1
            create_rule(
                version,
                ExamRuleTypeChoices.VISIBILITY,
                "field",
                field.id,
                {"patient_sex": ["female"]},
                {"visible": True},
                sort_order,
            )
            created += 1

    if sheet_name == "URINE- Clinical Microscopy":
        pregnancy_field = fields_by_key.get("microscopic_finding_pregnancy_test") or fields_by_key.get("pregnancy_test")
        if pregnancy_field:
            pregnancy_options = [key for key, option in options_by_key.items() if "pregnancy" in option.option_label.lower()]
            if pregnancy_options:
                sort_order += 1
                create_rule(
                    version,
                    ExamRuleTypeChoices.VISIBILITY,
                    "field",
                    pregnancy_field.id,
                    {"exam_option_keys": pregnancy_options},
                    {"visible": True},
                    sort_order,
                )
                created += 1

    if sheet_name == "OGTT - Blood Chemistry":
        option_map = {
            "50G ORAL GLUCOSE TOLERANCE": "50g ogtt",
            "75G ORAL GLUCOSE TOLERANCE": "75g ogtt",
            "100G ORAL GLUCOSE TOLERANCE": "100g ogtt",
        }
        for section_key, section in sections_by_key.items():
            label = normalize_text(section.section_label).upper()
            matched_option_keys = [
                key for key, option in options_by_key.items()
                if option_map.get(label, "").lower() == option.option_label.lower()
            ]
            if matched_option_keys:
                sort_order += 1
                create_rule(
                    version,
                    ExamRuleTypeChoices.VISIBILITY,
                    "section",
                    section.id,
                    {"exam_option_keys": matched_option_keys},
                    {"visible": True},
                    sort_order,
                )
                created += 1

    if sheet_name == "PROTIME, APTT - Hematology":
        mappings = {
            "pro_time": ["protime", "protime_aptt"],
            "aptt": ["aptt", "protime_aptt"],
        }
        for section_key, allowed in mappings.items():
            section = sections_by_key.get(section_key)
            if not section:
                continue
            matched_option_keys = [key for key in options_by_key if key in allowed]
            if matched_option_keys:
                sort_order += 1
                create_rule(
                    version,
                    ExamRuleTypeChoices.VISIBILITY,
                    "section",
                    section.id,
                    {"exam_option_keys": matched_option_keys},
                    {"visible": True},
                    sort_order,
                )
                created += 1

    if sheet_name == "COVID 19 ANTIGEN (RAPID TEST) -":
        attachment_field = fields_by_key.get("result_image")
        if attachment_field:
            sort_order += 1
            create_rule(
                version,
                ExamRuleTypeChoices.REQUIREMENT,
                "field",
                attachment_field.id,
                {"item_status": ["released"]},
                {"required": True},
                sort_order,
            )
            created += 1

    return created


def default_render_profile(sheet_name, has_sections, has_reference_ranges):
    layout_type = RenderLayoutTypeChoices.LABEL_VALUE_LIST
    if has_sections:
        layout_type = RenderLayoutTypeChoices.SECTIONED_REPORT
    elif has_reference_ranges:
        layout_type = RenderLayoutTypeChoices.RESULT_TABLE

    config = {
        "sheet_name": sheet_name,
        "show_reference_ranges": has_reference_ranges,
        "show_units": True,
        "render_variant": "generic",
    }

    if sheet_name == "ABG - Blood Gas Analysis":
        config.update(
            {
                "render_variant": "abg_compact",
                "left_section_key": "blood_gas_value_abg",
                "right_section_keys": [
                    "calculated_values_oximetry",
                    "calculated_values_acid_base_status",
                ],
            }
        )
    elif sheet_name == "BBANK - Blood Bank":
        config.update(
            {
                "render_variant": "bbank_crossmatch",
                "show_reference_ranges": False,
                "general_field_keys": [
                    "patients_blood_type",
                    "blood_component",
                    "donors_blood_type",
                    "source_of_blood",
                    "serial_number",
                    "date_extracted",
                    "date_expiry",
                ],
                "crossmatch_section_key": "type_of_crossmatching",
                "crossmatch_result_field_keys": [
                    "type_of_crossmatching_immediate_spin_saline_phase",
                    "type_of_crossmatching_albumin_phase_37_c",
                    "type_of_crossmatching_anti_human_globilin_phase",
                ],
                "remarks_field_key": "type_of_crossmatching_remarks",
                "vital_signs_field_key": "type_of_crossmatching_vital_signs",
                "release_field_keys": [
                    "type_of_crossmatching_released_by",
                    "type_of_crossmatching_released_to",
                    "type_of_crossmatching_datetime",
                ],
            }
        )
    elif sheet_name == "SEROLOGY":
        config.update(
            {
                "render_variant": "serology_panel",
                "show_reference_ranges": False,
                "option_to_section_keys": {
                    "dengue_test": "dengue_test",
                    "tyhpidot": "typhidot",
                },
                "option_to_field_keys": {
                    "hbsag_screening": ["hbsag_screening"],
                    "vdrl": ["vdrl"],
                    "anti_hcv": ["anti_hcv"],
                    "aso_titer": ["aso_titer"],
                },
            }
        )
    elif sheet_name == "CARDIACI - Serology":
        config.update(
            {
                "render_variant": "serology_panel",
                "show_reference_ranges": True,
                "option_to_field_keys": {
                    "ck_mb_tni_bnp": ["ck_mb", "troponin_i", "bnp"],
                },
            }
        )
    elif sheet_name == "OGTT - Blood Chemistry":
        config.update(
            {
                "render_variant": "ogtt_timeline",
                "show_reference_ranges": True,
                "option_to_section_keys": {
                    "50g_ogtt": "50g_oral_glucose_tolerance",
                    "75g_ogtt": "75g_oral_glucose_tolerance",
                    "100g_ogtt": "100g_oral_glucose_tolerance",
                },
                "option_to_field_keys": {
                    "2_hour_postprandial": ["2_hours_post_prandial"],
                    "50g_oral_glucose_challenge": ["50_g_oral_glucose_challenge"],
                },
            }
        )
    elif sheet_name == "PROTIME, APTT - Hematology":
        config.update(
            {
                "render_variant": "coagulation_panel",
                "show_reference_ranges": True,
                "option_to_sections": {
                    "protime": ["pro_time"],
                    "aptt": ["aptt"],
                    "protime_aptt": ["pro_time", "aptt"],
                },
            }
        )
    elif sheet_name == "HEMATOLOGY":
        config.update(
            {
                "render_variant": "hematology_panel",
                "option_to_panels": {
                    "cbc_platelet_count_blood_typing": [
                        {"title": "Cell Counts", "keys": ["rbc_count", "wbc_count", "hemoglobin", "hematocrit", "platelet_count"]},
                        {"title": "Differential Count", "keys": ["segmenters", "lymphocytes", "monocytes", "eosinophils", "stab"]},
                        {"title": "Blood Typing", "keys": ["blood_typing"]},
                    ],
                    "hgb_hct_platelet_count": [
                        {"title": "Core Results", "keys": ["hemoglobin", "hematocrit", "platelet_count"]},
                    ],
                    "blood_typing": [
                        {"title": "Blood Typing", "keys": ["blood_typing"]},
                    ],
                    "cbc_platelet_count_esr": [
                        {"title": "Cell Counts", "keys": ["rbc_count", "wbc_count", "hemoglobin", "hematocrit", "platelet_count"]},
                        {"title": "Differential Count", "keys": ["segmenters", "lymphocytes", "monocytes", "eosinophils", "stab"]},
                        {"title": "E.S.R.", "keys": ["esr"]},
                    ],
                    "cbc_platelet_count": [
                        {"title": "Cell Counts", "keys": ["rbc_count", "wbc_count", "hemoglobin", "hematocrit", "platelet_count"]},
                        {"title": "Differential Count", "keys": ["segmenters", "lymphocytes", "monocytes", "eosinophils", "stab"]},
                    ],
                },
                "sex_specific_field_map": {
                    "rbc_count": {
                        "male": "rbc_count_m",
                        "female": "rbc_count_f",
                        "label": "RBC COUNT",
                    },
                    "hemoglobin": {
                        "male": "hemoglobin_m",
                        "female": "hemoglobin_f",
                        "label": "HEMOGLOBIN",
                    },
                    "hematocrit": {
                        "male": "hematocrit_m",
                        "female": "hematocrit_f",
                        "label": "HEMATOCRIT",
                    },
                    "esr": {
                        "male": "esr_m",
                        "female": "esr_f",
                        "label": "E.S.R.",
                    },
                },
            }
        )
    elif sheet_name in {"BCMALE - Blood Chemistry", "BCFEMALE - Blood Chemistry"}:
        config.update(
            {
                "render_variant": "chemistry_panel",
                "show_reference_ranges": True,
                "panel_groups": [
                    {
                        "title": "Glucose Studies",
                        "keys": ["fasting_blood_sugar", "random_blood_sugar", "hgt"],
                    },
                    {
                        "title": "Renal Function",
                        "keys": ["blood_urea_nitrogen", "creatinine", "blood_uric_acid"],
                    },
                    {
                        "title": "Electrolytes and Minerals",
                        "keys": ["sodium", "potassium", "chloride", "ionized_calcium"],
                    },
                    {
                        "title": "Lipid Profile",
                        "keys": [
                            "cholesterol",
                            "triglyceride",
                            "hdl_cholesterol",
                            "ldl_cholesterol",
                            "vldl_cholesterol",
                        ],
                    },
                    {
                        "title": "Liver Enzymes",
                        "keys": ["sgotast", "sgptalt"],
                    },
                    {
                        "title": "Additional Result",
                        "keys": ["others"],
                    },
                ],
            }
        )
    elif sheet_name == "SEMEN - Clinical Microscopy":
        config.update(
            {
                "render_variant": "semen_analysis",
                "show_reference_ranges": False,
                "sample_field_keys": [
                    "time_collected",
                    "time_received",
                    "total_volume",
                    "liquefaction_time",
                ],
                "section_keys": [
                    "motility",
                    "morphology",
                    "sperm_count",
                    "others",
                ],
            }
        )
    elif sheet_name == "URINE- Clinical Microscopy":
        config.update(
            {
                "render_variant": "microscopy_sections",
                "show_reference_ranges": False,
                "option_to_sections": {
                    "urinalysis": [
                        "macroscopic_finding",
                        "clinical_finding",
                        "microscopic_finding",
                    ],
                    "urinalysis_urine_ketone": [
                        "macroscopic_finding",
                        "clinical_finding",
                        "microscopic_finding",
                    ],
                    "urinalysis_pregnancy_test": [
                        "macroscopic_finding",
                        "clinical_finding",
                        "microscopic_finding",
                    ],
                    "urinalysis_pregnancy_test_blood": [
                        "macroscopic_finding",
                        "clinical_finding",
                        "microscopic_finding",
                    ],
                },
                "option_to_field_keys": {
                    "pregnancy_test": ["microscopic_finding_pregnancy_test"],
                    "pregnancy_test_blood": ["microscopic_finding_pregnancy_test"],
                },
                "option_to_excluded_field_keys": {
                    "urinalysis": ["microscopic_finding_pregnancy_test"],
                    "urinalysis_urine_ketone": ["microscopic_finding_pregnancy_test"],
                },
            }
        )
    elif sheet_name == "FECALYSIS - Clinical Microscopy":
        config.update(
            {
                "render_variant": "microscopy_sections",
                "show_reference_ranges": False,
                "option_to_sections": {
                    "fecalysis": [
                        "macroscopic_finding",
                        "microscopic_finding",
                        "microscopic_finding_2",
                    ],
                    "fecalysis_fobt": [
                        "macroscopic_finding",
                        "microscopic_finding",
                        "microscopic_finding_2",
                    ],
                },
                "option_to_field_keys": {
                    "fobt": ["macroscopic_finding_fecal_occult_blood"],
                },
                "option_to_excluded_field_keys": {
                    "fecalysis": ["macroscopic_finding_fecal_occult_blood"],
                },
            }
        )
    elif sheet_name == "MICROBIOLOGY":
        config.update(
            {
                "render_variant": "single_result_focus",
                "show_reference_ranges": False,
                "field_keys": ["result"],
            }
        )

    return layout_type, config


@transaction.atomic
def import_workbook(
    workbook_path,
    publish=True,
    archive_old=True,
    force=False,
    importer_signature_version=IMPORTER_SIGNATURE_VERSION,
):
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, data_only=True)
    stats = ImportStats()

    for ws in wb.worksheets:
        meta = make_meta(ws.title)
        payload = sheet_payload(ws)
        payload_rows = payload["rows"]
        raw_options_col = payload["columns"]["raw_options"]
        sheet_notes = payload.get("sheet_notes", [])
        source_reference = build_source_reference(
            sheet_name=ws.title,
            payload=payload,
            importer_signature_version=importer_signature_version,
        )
        unscoped_fields = {normalize_text(value) for value in UNSCOPED_FIELDS_BY_SHEET.get(ws.title, set())}

        exam_definition, created_definition = ExamDefinition.objects.get_or_create(
            exam_code=meta["exam_code"],
            defaults={
                "exam_name": meta["exam_name"],
                "category": meta["category"],
                "description": ws.title,
                "active": True,
            },
        )
        if created_definition:
            stats.created_definitions += 1
        else:
            exam_definition.exam_name = meta["exam_name"]
            exam_definition.category = meta["category"]
            exam_definition.description = ws.title
            exam_definition.active = True
            exam_definition.save(update_fields=["exam_name", "category", "description", "active", "updated_at"])

        latest_published = exam_definition.versions.filter(version_status=ExamVersionStatusChoices.PUBLISHED).order_by("-version_no").first()
        if not force and latest_published and latest_published.source_reference == source_reference:
            stats.skipped_versions += 1
            continue

        if archive_old:
            exam_definition.versions.filter(version_status=ExamVersionStatusChoices.PUBLISHED).update(
                version_status=ExamVersionStatusChoices.ARCHIVED
            )

        next_version_no = (exam_definition.versions.order_by("-version_no").values_list("version_no", flat=True).first() or 0) + 1
        version = ExamDefinitionVersion.objects.create(
            exam_definition=exam_definition,
            version_no=next_version_no,
            version_status=ExamVersionStatusChoices.PUBLISHED if publish else ExamVersionStatusChoices.DRAFT,
            source_type="xlsx_import",
            source_reference=source_reference,
            published_at=timezone.now() if publish else None,
            change_summary=f"Imported from workbook sheet: {ws.title}",
        )
        stats.created_versions += 1

        options_by_key = {}
        sections_by_key = {}
        fields_by_key = {}
        section_occurrences = {}
        current_section = None
        section_sort_order = 0
        field_sort_order = 0

        exam_option_lines = split_multiline(get_cell_value(ws, 7, raw_options_col))
        for position, option_label in enumerate(exam_option_lines, start=1):
            option = ExamOption.objects.create(
                exam_version=version,
                option_key=make_key(option_label),
                option_label=option_label,
                sort_order=position,
                active=True,
            )
            options_by_key[option.option_key] = option
            stats.created_options += 1

        for row in payload_rows:
            field_label = row["field"]
            input_type = row["input_type"]
            raw_options = row["raw_options"]
            reference = row["reference"]
            notes = row["notes"]

            if not field_label or field_label == "Field":
                continue

            if field_label in PATIENT_FIELDS or field_label in SIGNATORY_FIELDS:
                continue

            if is_internal_note_row(field_label):
                continue

            if not input_type:
                if field_label in SKIP_SECTION_LABELS:
                    current_section = None
                    continue

                base_section_key = make_key(field_label)
                occurrence = section_occurrences.get(base_section_key, 0) + 1
                section_occurrences[base_section_key] = occurrence
                section_key = base_section_key if occurrence == 1 else f"{base_section_key}_{occurrence}"
                section_sort_order += 1
                current_section = ExamSection.objects.create(
                    exam_version=version,
                    section_key=section_key,
                    section_label=field_label,
                    sort_order=section_sort_order,
                    active=True,
                )
                sections_by_key[section_key] = current_section
                stats.created_sections += 1
                continue

            effective_section = current_section
            normalized_field_label = normalize_text(field_label)
            if normalized_field_label in unscoped_fields:
                effective_section = None

            field_input_type, data_type = infer_field_types(ws.title, field_label, input_type, raw_options, reference)
            field_key = make_field_key(effective_section.section_key if effective_section else "", field_label)
            unit = extract_unit(input_type, raw_options)
            field_config = build_field_config(field_input_type, raw_options, internal_note=notes)
            field_sort_order += 1

            exam_field = ExamField.objects.create(
                exam_version=version,
                section=effective_section,
                field_key=field_key,
                field_label=field_label,
                input_type=field_input_type,
                data_type=data_type,
                unit=unit,
                required=False,
                sort_order=field_sort_order,
                default_value="",
                help_text="",
                placeholder_text="",
                reference_text=reference,
                config_json=field_config,
                supports_attachment=field_input_type == ExamFieldInputTypeChoices.ATTACHMENT,
                active=True,
            )
            fields_by_key[field_key] = exam_field
            stats.created_fields += 1

            if field_input_type == ExamFieldInputTypeChoices.SELECT:
                for position, option_label in enumerate(split_multiline(raw_options), start=1):
                    ExamFieldSelectOption.objects.create(
                        field=exam_field,
                        option_value=make_key(option_label),
                        option_label=option_label,
                        sort_order=position,
                        active=True,
                    )

            parsed_range = parse_reference_range(reference)
            if parsed_range:
                ExamFieldReferenceRange.objects.create(
                    field=exam_field,
                    option_scope=None,
                    sex_scope="",
                    range_type=parsed_range["range_type"],
                    min_numeric=parsed_range["min_numeric"],
                    max_numeric=parsed_range["max_numeric"],
                    reference_text=parsed_range["reference_text"],
                    abnormal_rule=parsed_range["abnormal_rule"],
                    sort_order=1,
                )
                stats.created_ranges += 1

        if ws.title == "COVID 19 ANTIGEN (RAPID TEST) -":
            field_sort_order += 1
            covid_note = sheet_notes[0] if sheet_notes else "kelangan may picture sa result"
            attachment_field = ExamField.objects.create(
                exam_version=version,
                section=None,
                field_key="result_image",
                field_label="Result Image",
                input_type=ExamFieldInputTypeChoices.ATTACHMENT,
                data_type=ExamFieldDataTypeChoices.STRING,
                unit="",
                required=False,
                sort_order=field_sort_order,
                default_value="",
                help_text="",
                placeholder_text="",
                reference_text="",
                config_json={"internal_note": f"Imported from workbook note: {covid_note}"},
                supports_attachment=True,
                active=True,
            )
            fields_by_key[attachment_field.field_key] = attachment_field
            stats.created_fields += 1

        stats.created_rules += build_safe_rules(ws.title, version, options_by_key, sections_by_key, fields_by_key)

        has_sections = bool(sections_by_key)
        has_reference_ranges = ExamFieldReferenceRange.objects.filter(field__exam_version=version).exists()
        layout_type, render_config = default_render_profile(ws.title, has_sections, has_reference_ranges)
        ExamRenderProfile.objects.create(
            exam_version=version,
            layout_type=layout_type,
            config_json=render_config,
            active=True,
        )

    return stats
